# (pip install openpyxl)
from openpyxl import Workbook

inStr=""
#count=1
count=13001

wb=Workbook()

while True:
    # 파일 입력
    folderName='S'+'0'*(8-len(str(count)))+str(count)
    #if folderName=='S00000090': # S00000090이 없음
        #continue
    filePath='/Users/최웅지/Downloads/고객 응대 음성 데이터/Training/[라벨]D50_J01/'+folderName+'/0001.txt'
    #f=open(filePath, 'r', encoding="utf-8")
    try:
        f=open(filePath, 'r', encoding="utf-8")
    except FileNotFoundError:
        count+=1
        continue
    else:
        inStr=f.read()

        # 엑셀 파일에 쓰기
        sheet1=wb.active
        sheet1.cell(count-13000,1,inStr)
        #sheet2=wb.create_sheet('result') # 새 시트 만들기
        #sheet2.title='example' # 시트 이름 바꾸기
        count+=1

        #if count==50:
        if count==20000:
            break

wb.save('/Users/최웅지/Documents/2021 여름/dataset.xlsx')
f.close()
